# -*- coding: utf-8 -*-
"""
Created on Wed Feb 10 06:38:58 2016

ToDo:
Doku
- test for last entry
- write data in table

Anweisung:
- zu messende bzw berechnende Größen übergeben bzw anfordern
- IDs
@author: Roman
"""

from PyQt4 import QtGui, QtCore
from openpyxl import load_workbook
from time import strftime
from shutil import copyfile
#import os
#print(strftime("%d.%m.%Y"))

class ProductionInfo(QtGui.QWidget):
    emitSoll = QtCore.pyqtSignal(str)
    emitTol = QtCore.pyqtSignal(str)
    emitProdIds = QtCore.pyqtSignal(str, str)
    
    emitGetIDs = QtCore.pyqtSignal()
    emitTestFbg = QtCore.pyqtSignal(int) # int - channel number
    emitClearIDs = QtCore.pyqtSignal()
    def __init__(self, *args):
        QtGui.QWidget.__init__(self, *args)
        
        self.setMaximumHeight(500)

        self.__paths = ['','','']
        self.__pathLabels = ['Produktionsplan', 'Produktionsdokumentation', 
                  'Spektrenordner']
                  
        self.proStepNb = []
        self.proShort = []
        self.proDiscription = []
        self.sollGreen = []
        self.sollArray = []
        self.proMeas = []
        self.proCond = []
        self.__log = None #Excel worksheet for data storage
        self.__logRow = 5 #start row, will set to actual row
        self.__proId = ''
        self.__year = 2016
        
        self.__step = 0
        
        self.__prodPlanNum = 0 # 0 - build, 1 - silicon, 2 - fiber termination
        
        self.__startString = [unicode('\tDrücken Sie START\n um die Sensorfertigung zu beginnen', 'utf-8'),
                              unicode('\tDrücken Sie START\n um mit dem Einbringen des Silikons zu beginnen', 'utf-8'),
                                unicode('\tDrücken Sie START\n um mit der Faserterminierung zu beginnen', 'utf-8')]
                                
        self.__planString = ['acc_Plan','acc_Silikon','acc_Terminierung', 'acc_QC']
        self.__planMessage = ['Sensoraufbau','Silikon','Faserterminierung', 'Endkontrolle']
                                
        #Silikon
        self.__prodID = []
        self.__fbgID = []
        self.__sensID = []
        self.__idRow = []
        
        self.info = QtGui.QTextEdit()
        self.info.setReadOnly(True)
        self.info.setFontPointSize(20)
        self.info.setAlignment(QtCore.Qt.AlignHCenter)
        self.info.setText(self.__startString[0])
        self.info.setAlignment(QtCore.Qt.AlignLeft)
        
        hl = QtGui.QHBoxLayout(self)
        hl.addWidget(self.info)
        self.buttons = Buttons()
        hl.addWidget(self.buttons)
        
        self.loadSettings()
        self.buttons.emitStop.connect(self.productionSequenzCancel)
        
    def changeProdPlan(self, planNum = 0):
        if self.__step and self.__step < int(self.proStepNb[-1]-1):
            reply = QtGui.QMessageBox.question(self, 'Achtung',
                                            unicode("""Dies führt zum Abbruch der aktuellen Produktionssequenz!
                                            \n\n Wollen Sie wirklich fortfahren""", 'utf-8'), 
                                            QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
        
            if reply == QtGui.QMessageBox.Yes:
                self.productionSequenzNew(planNum)
                return 1
            else:
                return 0
        else:
            self.productionSequenzNew(planNum)
            return 1
    
    def generateIDs(self):
        if self.__proId:
            id = self.__proId.split('-')
            num = int(id[-1])+1
            proID = str(id[0])+str('-')+str(num).zfill(4)
        else:
            proID = str(self.__year)+str('-')+str(1).zfill(4)
        if self.__logRow > 5:
            sensorID = self.__log['C'+str(self.__logRow-1)].value
        else: 
            sensorID = 0
        num = int(sensorID)+1
        sensorID = str(num).zfill(4)
        self.emitProdIds.emit(proID, sensorID)
        
    def getIDbyIndex(self, index = -1):
        if not index == -1:
            self.__logRow = self.__idRow[index]
        return self.__prodID[index], self.__fbgID[index], self.__sensID[index]
        
    def getProCondition(self):
        return self.proCond[self.__step]   
        
    def getProMeas(self):
        return self.proMeas[self.__step]
        
    def getProdPlanNum(self):
        return self.__prodPlanNum
    
    def getProStep(self):
        return self.__step
        
    def getSensorIDs(self):
        return self.__sensID
        
    def getTolaranz(self):
        return self.__TolArray[self.sollArray[self.__step]-1]
        
    def loadProductionPlan(self):
        print('Lade Arbeitsplan %s' % self.__planMessage[self.__prodPlanNum])
        if not self.__paths[0]:
            QtGui.QMessageBox.critical(self, 'Keine Produktionsdatei','Kein Pfad angegeben! Bitte in Optionen nachtragen.')
            return 0
        self.proStepNb = []
        self.proShort = []
        self.proDiscription = []
        self.sollArray = []
        self.proMeas = []
        self.proCond = []
        filename=self.__paths[0]
        print(filename)
        wb = load_workbook(filename)
        
        ws = wb[self.__planString[self.__prodPlanNum]]
        content = True
        row = 2
        self.__prodID = []; self.__fbgID = []; self.__sensID = []; self.__idRow = []
        
        while content:
            cell = ws['C'+str(row)].value
            #print cell
            if cell: 
                content = True
            else:
                content = False
                break
            self.proStepNb.append(cell)
            cell = ws['D'+str(row)].value
            self.proShort.append(self.testCell(cell))
            cell = ws['F'+str(row)].value
            self.proDiscription.append(self.testCell(cell))
            cell = ws['G'+str(row)].value
            self.sollArray.append(int(self.testCellNum(cell)))
            cell = ws['I'+str(row)].value
            self.proMeas.append(self.testCell(cell))
            cell = ws['J'+str(row)].value
            self.proCond.append(self.testCell(cell))
            
            row += 1
        print('Arbeitsplan %s wurde geladen' % self.__planMessage[self.__prodPlanNum])
        
    def loadProductionTable(self):
        self.loadProductionPlan()
        
        print('Lade Log-Datei')
        self.wb = load_workbook(filename=self.__paths[1])
        self.__log = self.wb['acc_Produktion']
        self.__logRow = 5        
        content = True
        
        while content:
            cell = self.__log['A'+str(self.__logRow)].value
            #print('Zeile ', str(self.__logRow),': ', cell)
            if cell:
                content = True
                self.__proId = cell
                self.__prodID.append(self.testCell(cell))
                cell = self.__log['B'+str(self.__logRow)].value
                self.__fbgID.append(self.testCell(cell))
                cell = self.__log['C'+str(self.__logRow)].value
                self.__sensID.append(self.testCell(cell))
                self.__logRow += 1
            else:
                content = False
            
           
        #self.generateIDs()
        print('Log-Datei wurde geladen')
        return 1
        
    def loadProductionLogSi(self, _file = None):
        self.loadProductionPlan()
        
        print('Lade Log-Datei')
        self.wb = load_workbook(filename=self.__paths[1])
        self.__log = self.wb['acc_Produktion']
        self.__prodID = []; self.__fbgID = []; self.__sensID = []; self.__idRow = []
        self.__logRow = 5
        content = True
        
        while content:
            cell = self.__log['A'+str(self.__logRow)].value
            if cell:
                content = True
                if not self.__log['L'+str(self.__logRow)].value:
                    self.__prodID.append(self.testCell(cell))
                    cell = self.__log['B'+str(self.__logRow)].value
                    self.__fbgID.append(self.testCell(cell))
                    cell = self.__log['C'+str(self.__logRow)].value
                    self.__sensID.append(self.testCell(cell))
                    self.__idRow.append(self.__logRow)
                self.__logRow +=1
                
            else:
                content = False
                
        if len(self.__sensID)>1:
            self.__sensID, self.__prodID, self.__fbgID, self.__idRow = zip(*sorted(zip(self.__sensID, self.__prodID, self.__fbgID, self.__idRow)))
        print('Log-Datei wurde geladen')
        
    def loadProductionLogTerm(self, _file = None):
        self.loadProductionPlan()
        
        print('Lade Log-Datei')
        self.wb = load_workbook(filename=self.__paths[1])
        self.__log = self.wb['acc_Produktion']
        self.__prodID = []; self.__fbgID = []; self.__sensID = []; self.__idRow = []
        self.__logRow = 5
        content = True
        
        while content:
            cell = self.__log['A'+str(self.__logRow)].value
            if cell:
                content = True
                if not self.__log['O'+str(self.__logRow)].value:
                    self.__prodID.append(self.testCell(cell))
                    cell = self.__log['B'+str(self.__logRow)].value
                    self.__fbgID.append(self.testCell(cell))
                    cell = self.__log['C'+str(self.__logRow)].value
                    self.__sensID.append(self.testCell(cell))
                    self.__idRow.append(self.__logRow)
                self.__logRow +=1
                
            else:
                content = False
                
        if len(self.__sensID)>1:
            self.__sensID, self.__prodID, self.__fbgID, self.__idRow = zip(*sorted(zip(self.__sensID, self.__prodID, self.__fbgID, self.__idRow)))
        print('Log-Datei wurde geladen')
        
    def loadSettings(self):
        print('Lade Einstellungen Produktionsverlauf')
        self.__TargetArray = []
        self.__TolArray = []
        settings = QtCore.QSettings('test.ini',QtCore.QSettings.IniFormat)
        settings.beginGroup('Produktion')
        self.__TargetArray.append(settings.value('VorspannGrob').toFloat()[0])
        self.__TolArray.append(settings.value('VorGrobTol').toFloat()[0])
        self.__TargetArray.append(settings.value('VorspannFein').toFloat()[0])
        self.__TolArray.append(settings.value('VorFeinTol').toFloat()[0])
        self.__TargetArray.append(settings.value('SiVerklebung').toFloat()[0])
        self.__TolArray.append(settings.value('DeltaSiVerklebung').toFloat()[0])
        self.__TargetArray.append(settings.value('returnLoss').toInt()[0])
        self.__TolArray.append(settings.value('DeltaReturnLoss').toInt()[0])
        settings.endGroup()
        settings.beginGroup('Dateipfade')
        for i in range(3):
            s = str(settings.value(self.__pathLabels[i]).toString())
            self.__paths[i] = s
        settings.endGroup()
        print('Einstellungen Produktionsverlauf wurden geladen')
        
    def productionSequenzBack(self):
        self.__step -= 2
        self.nextProductionStep()
        
    def productionSequenzCancel(self, cancel = True):
        print('Cancel: ', cancel)
        if cancel:
            reply = QtGui.QMessageBox.question(self, 'Achtung',
                                        unicode("""Es wird eine neue Produktionssequenz gestartet!
                                        \n\n Wollen Sie wirklich fortfahren""", 'utf-8'), 
                                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
    
            if reply == QtGui.QMessageBox.Yes:
                print('Neu')
                if self.__prodPlanNum == 0:
                    print('clear row', self.__logRow)
                    self.__log['A'+str(self.__logRow)].value = str('')
                elif self.__prodPlanNum == 1:
                    self.__log['J'+str(self.__logRow)].value = str('')
                elif self.__prodPlanNum == 2:
                    self.__log['O'+str(self.__logRow)].value = str('')
                self.wb.save(self.__paths[1])
                self.emitClearIDs.emit()
                
                self.startProduction()
            else:
                return 0
        
        else:        
            self.emitClearIDs.emit()
            fname = self.__paths[1].split('.')
            fn = fname[0] + "_save." + fname[1]
            copyfile(self.__paths[1], fn)
            self.startProduction()
                
    def productionSequenzNew(self, planNum):
        self.__step = 0
        self.__prodPlanNum = planNum
        self.info.setText(self.__startString[planNum])
        self.buttons.startButton.setText('Start')
        self.buttons.stopButton.setEnabled(False)
        self.buttons.backButton.setEnabled(False)
        
    def setIDs(self, pro, fbg, sensor):
        self.__log['A'+str(self.__logRow)].value = str(pro)
        self.__log['B'+str(self.__logRow)].value = str(fbg)
        self.__log['C'+str(self.__logRow)].value = str(sensor)
        self.wb.save(self.__paths[1])
        
        return 1
        
    def setDate(self):
        colDate = {0:'D', 1:'L',2:'O'}
        self.__log[colDate[self.__prodPlanNum]+str(self.__logRow)].value = strftime("%H:%M-%d.%m.%Y")
        self.wb.save(self.__paths[1])
        
    def setPeakWavelength(self, wavelength):
        if self.__prodPlanNum == 0:
            colPeakWL = {3:'E', 6:'J'}
        elif self.__prodPlanNum == 1:
            colPeakWL = {2:'M', 4:'N'}
        elif self.__prodPlanNum == 2:
            colPeakWL = {2:'P'}
        self.__log[colPeakWL[self.proStepNb[self.__step]]+str(self.__logRow)].value = float(wavelength)
        self.wb.save(self.__paths[1])
            
    def setSpecFile(self, fname):
        if self.proStepNb[self.__step] == 3:
            col = 'G'
            
        self.__log[col+str(self.__logRow)].value = str(fname)
        self.wb.save(self.__paths[1])
    
    def setTemp(self, temp):
        col = {3:'F',6:'K'}
        temp = temp.split(' ')
        self.__log[col[self.proStepNb[self.__step]]+str(self.__logRow)].value = float(temp[0])
        self.wb.save(self.__paths[1])
    
    def setFWHM(self, fwhm):
        if self.proStepNb[self.__step] == 3:
            col = 'H'
        
        self.__log[col+str(self.__logRow)].value = float(fwhm)
        self.wb.save(self.__paths[1])
    
    def setAsymmetrie(self,asym):
        if self.proStepNb[self.__step] == 3:
            col = 'I'
             
        self.__log[col+str(self.__logRow)].value = float(asym)
        self.wb.save(self.__paths[1])
        
    def testID(self, ID, val):
        if ID == 0:
            if val in self.__prodID:
                return 0
            else:
                return 1
        elif ID == 1:
            if val in self.__fbgID:
                return 0
            else:
                return 1
        elif ID == 2:
            if val in self.__sensID:
                return 0
            else:
                return 1        
        else:
            return 1
            
    def testCell(self,cell): 
        if cell:
             cell = unicode(cell)
        else:
             cell = unicode('')
        return cell
        
    def testCellNum(self,cell): 
        if cell:
             return cell
        else:
             cell = 0.
        return cell

    def makeProTxt(self, num=0):
        step = str(self.proStepNb[num]) + '/'+str(self.proStepNb[-1])+': ' + self.proShort[num]
        discription = self.proDiscription[num]
        goal = 'Zielwert: '
        strFormat = "{0:.3f}"
        if self.__prodPlanNum == 2:
            strFormat = "{0:.1f}"
        if self.sollArray[num]:
            goal = goal + str(strFormat.format(self.__TargetArray[self.sollArray[num]-1])) + u' \u00B1 ' + str(strFormat.format(self.__TolArray[self.sollArray[num]-1]))
            self.emitSoll.emit(str(strFormat.format(self.__TargetArray[self.sollArray[num]-1])))
        else:
            goal =''
            self.emitSoll.emit('')
        
        self.info.clear()
        self.info.setFontUnderline(True)
        self.info.setFontPointSize(12)
        self.info.append(step)
        self.info.setFontUnderline(False)
        self.info.setFontPointSize(11)
        self.info.append('')
        self.info.append(goal)
        self.info.append('')
        
        self.info.append(discription)
        
    def nextProductionStep(self):
        self.__step += 1
        print('Step: ', self.__step)
        if self.__step < self.proStepNb[-1]:
            self.makeProTxt(self.__step)
            #self.chan1SollLabel.setText()
        self.buttons.backButton.setEnabled(self.__step)
        self.buttons.stopButton.setEnabled(self.__step)
        
    def startProduction(self):
        if self.__prodPlanNum == 0:
            if not self.loadProductionTable():
                return 0
            self.generateIDs()
        elif self.__prodPlanNum == 1:
            self.loadProductionLogSi()
        elif self.__prodPlanNum == 2:   
            self.loadProductionLogTerm()
        
        self.__step = 0
        self.makeProTxt(self.__step)
        self.buttons.startButton.setText('Weiter')
        self.buttons.stopButton.setEnabled(False)
        self.buttons.backButton.setEnabled(False)
        
        
class Buttons(QtGui.QWidget):
    emitStop = QtCore.pyqtSignal(int)
    def __init__(self,*args):
        QtGui.QWidget.__init__(self,*args)
        
        style = "font-size: 20px; font-weight: bold"
        self.startButton = QtGui.QPushButton(text='Start')
        self.startButton.setStyleSheet(style)
        self.backButton = QtGui.QPushButton(text=unicode('Zurück', 'utf-8'))
        self.backButton.setStyleSheet(style)
        self.backButton.setEnabled(False)
        self.stopButton = QtGui.QPushButton(text='Abbruch')
        self.stopButton.setStyleSheet(style)
        self.stopButton.setEnabled(False)
        self.stopButton.clicked.connect(self.stopClicked)
        
        bl = QtGui.QVBoxLayout()
        bl.addWidget(self.startButton)
        bl.addWidget(self.backButton)
        bl.addWidget(self.stopButton)
        self.setLayout(bl)
        
    def stopClicked(self):
        self.emitStop.emit(1)
        
    def setEnabled(self, val):
        self.startButton.setEnabled(val)
        self.backButton.setEnabled(val)
        self.stopButton.setEnabled(val)
        
